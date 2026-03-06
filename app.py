import streamlit as st
import requests
import json
import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="⚽ BetSlip AI",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  CUSTOM CSS  — dark football stadium aesthetic
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=DM+Sans:wght@300;400;500;600;700&display=swap');

:root {
    --green-dark:  #0d3320;
    --green-mid:   #1a5c2e;
    --green-main:  #2e8b57;
    --green-light: #d6f0e0;
    --gold:        #f5a623;
    --gold-dark:   #c8960c;
    --red:         #c0392b;
    --red-light:   #fdecea;
    --bg:          #0a1628;
    --bg2:         #0f1f35;
    --surface:     rgba(255,255,255,0.05);
    --border:      rgba(255,255,255,0.08);
    --text:        #e8f0fe;
    --text-muted:  #7a8fa6;
}

html, body, [data-testid="stAppViewContainer"] {
    background: linear-gradient(160deg, #0a1628 0%, #0d2137 60%, #0a1628 100%) !important;
    color: var(--text) !important;
}

[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d3320 0%, #071a10 100%) !important;
    border-right: 2px solid var(--green-main) !important;
}

[data-testid="stSidebar"] * { color: #e8f0fe !important; }

h1, h2, h3 { font-family: 'Bebas Neue', sans-serif !important; letter-spacing: 2px; }
p, div, label, span { font-family: 'DM Sans', sans-serif !important; }

.stButton > button {
    background: linear-gradient(135deg, var(--green-mid), var(--green-main)) !important;
    color: white !important;
    font-family: 'Bebas Neue', sans-serif !important;
    font-size: 1.1rem !important;
    letter-spacing: 2px !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.6rem 2rem !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 20px rgba(46,139,87,0.3) !important;
    width: 100% !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, var(--green-main), #3daa6e) !important;
    box-shadow: 0 6px 28px rgba(46,139,87,0.5) !important;
    transform: translateY(-1px) !important;
}

.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: rgba(0,0,0,0.4) !important;
    border: 1px solid var(--green-main) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}

.stDateInput > div > div > input {
    background: rgba(0,0,0,0.4) !important;
    border: 1px solid var(--green-main) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}

.stTextInput > div > div > input {
    background: rgba(0,0,0,0.4) !important;
    border: 1px solid var(--green-main) !important;
    color: var(--text) !important;
    border-radius: 8px !important;
}

.stCheckbox > label { color: var(--text) !important; font-size: 0.95rem !important; }

div[data-testid="stMetric"] {
    background: rgba(255,255,255,0.04) !important;
    border: 1px solid var(--border) !important;
    border-radius: 10px !important;
    padding: 12px !important;
}
div[data-testid="stMetricValue"] { color: var(--gold) !important; font-family: 'Bebas Neue', sans-serif !important; font-size: 1.6rem !important; }
div[data-testid="stMetricLabel"] { color: var(--text-muted) !important; font-size: 0.75rem !important; letter-spacing: 1px; text-transform: uppercase; }

.stTabs [data-baseweb="tab-list"] {
    background: rgba(0,0,0,0.3) !important;
    border-radius: 10px !important;
    padding: 4px !important;
    gap: 4px !important;
}
.stTabs [data-baseweb="tab"] {
    background: transparent !important;
    color: var(--text-muted) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    border-radius: 8px !important;
}
.stTabs [aria-selected="true"] {
    background: var(--green-mid) !important;
    color: white !important;
}

.stAlert { border-radius: 8px !important; }

div[data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #1a3a6e, #1e4d9a) !important;
    color: white !important;
    font-family: 'Bebas Neue', sans-serif !important;
    letter-spacing: 2px !important;
    font-size: 1rem !important;
    border: none !important;
    border-radius: 8px !important;
    box-shadow: 0 4px 16px rgba(26,58,110,0.4) !important;
}

hr { border-color: var(--border) !important; }

.slip-card {
    background: rgba(255,255,255,0.04);
    border-radius: 14px;
    padding: 20px;
    border: 1px solid rgba(255,255,255,0.08);
    margin-bottom: 16px;
}
.slip-header {
    background: linear-gradient(135deg, #1a5c2e, #2e8b57);
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 16px;
}
.insane-header {
    background: linear-gradient(135deg, #8b0000, #c0392b);
    border-radius: 10px;
    padding: 16px 20px;
    margin-bottom: 16px;
    border: 2px solid #f5a623;
}
.fixture-row {
    display: flex;
    align-items: center;
    padding: 10px 14px;
    border-radius: 8px;
    margin-bottom: 4px;
    font-family: 'DM Sans', sans-serif;
    font-size: 0.88rem;
}
.fixture-row:nth-child(odd)  { background: rgba(46,139,87,0.08); }
.fixture-row:nth-child(even) { background: rgba(255,255,255,0.03); }
.tag {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.5px;
}
.tag-green  { background: rgba(46,139,87,0.25); color: #7ecf9e; border: 1px solid #2e8b57; }
.tag-orange { background: rgba(230,126,34,0.25); color: #f0a060; border: 1px solid #e67e22; }
.tag-red    { background: rgba(192,57,43,0.25); color: #ff8a7a; border: 1px solid #c0392b; }
.tag-blue   { background: rgba(26,58,110,0.3);  color: #7aabea; border: 1px solid #1a3a6e; }
.tag-gold   { background: rgba(245,166,35,0.2); color: #f5a623; border: 1px solid #c8960c; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
LEAGUES = {
    "bundesliga":       ("🇩🇪", "Bundesliga"),
    "la_liga":          ("🇪🇸", "La Liga"),
    "serie_a":          ("🇮🇹", "Serie A"),
    "ligue_1":          ("🇫🇷", "Ligue 1"),
    "eredivisie":       ("🇳🇱", "Eredivisie"),
    "epl":              ("🏴", "Premier League"),
    "champions_league": ("🏆", "Champions League"),
}

SLIP_TYPES = {
    "fortress":   ("🏰", "Fortress",       "8 strongest favourites"),
    "goals":      ("⚽", "Goals Machine",  "Over goals & BTTS markets"),
    "protection": ("🛡️", "Protection",     "Draw No Bet & Double Chance"),
    "mixed":      ("🌍", "Mixed Best",     "Best pick per league"),
    "insane":     ("🤪", "INSANE",         "16 legs – 50,000x+ dream"),
    "bonus":      ("⭐", "Claude's Pick",  "Best balance selection"),
}

SLIP_ACCENTS = {
    "fortress":   ("#1a5c2e", "#d6f0e0", "tag-green"),
    "goals":      ("#e67e22", "#fef0e0", "tag-orange"),
    "protection": ("#1a3a6e", "#d6e8fa", "tag-blue"),
    "mixed":      ("#2e8b57", "#d6f0e0", "tag-green"),
    "bonus":      ("#c8960c", "#fff8dc", "tag-gold"),
    "insane":     ("#c0392b", "#fdecea", "tag-red"),
}

SYSTEM_PROMPT = """You are an elite football betting analyst with deep knowledge of all major European leagues.

The user gives you a date and leagues. Your job:
1. Generate realistic match fixtures for that date and leagues
2. Assign realistic win probabilities based on team quality, home advantage, recent form
3. Build the requested betting slips with sound reasoning

RULES:
- Only suggest realistic, plausible outcomes — no random upsets
- For INSANE slip: 14-18 legs, all realistic, win-to-nils vs weak sides, stacked markets, targeting 50,000x+
- Eredivisie = most goals per game in Europe (avg 3.2) — reflect this in goals picks
- Each regular slip = exactly 8 legs with detailed reasoning per leg
- Be specific: real team names, times, selections, probabilities

Respond ONLY with valid JSON — no markdown, no explanation, just the JSON:
{
  "date": "Sunday 8 March 2026",
  "summary": "Brief overview of the day's betting landscape",
  "fixtures": [
    {
      "time": "14:00",
      "league": "Ligue 1",
      "flag": "🇫🇷",
      "home": "Lens",
      "away": "Metz",
      "homeWinPct": 76,
      "drawPct": 14,
      "awayWinPct": 10,
      "bestBet": "Lens WIN"
    }
  ],
  "slips": [
    {
      "type": "fortress",
      "title": "THE FORTRESS",
      "subtitle": "8 strongest favourites across all leagues",
      "estimatedOdds": "20-35x",
      "risk": "Low-Medium",
      "riskColor": "green",
      "legs": 8,
      "selections": [
        {
          "num": 1,
          "match": "Lens vs Metz",
          "league": "Ligue 1",
          "flag": "🇫🇷",
          "selection": "Lens WIN",
          "prob": "76%",
          "reasoning": "Metz only 9% win prob — worst side in Ligue 1"
        }
      ],
      "analysis": [
        "Lens is the bedrock anchor at 76% probability.",
        "Key swing leg: check the lowest-probability selection carefully."
      ]
    }
  ]
}"""

# ─────────────────────────────────────────────────────────────────────────────
#  API CALL
# ─────────────────────────────────────────────────────────────────────────────
def call_claude(api_key, date_str, leagues, slip_types):
    league_names = [f"{LEAGUES[l][0]} {LEAGUES[l][1]}" for l in leagues if l in LEAGUES]
    slip_names   = [SLIP_TYPES[s][1] for s in slip_types if s in SLIP_TYPES]

    user_msg = f"""Generate betting slips for: {date_str}

Leagues: {', '.join(league_names)}
Slip types: {', '.join(slip_names)}

Notes:
- Include Eredivisie if selected — emphasise high-scoring Dutch football
- INSANE slip: 14-18 legs, win-to-nils on weak sides, targeting 50,000x+
- Include 12-15 fixtures in fixtures array (not more)
- Analysis per slip: 3-4 bullet points only (keep concise)
"""
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    body = {
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 16000,
        "system": SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_msg}],
    }
    resp = requests.post("https://api.anthropic.com/v1/messages",
                         headers=headers, json=body, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f"API error {resp.status_code}: {resp.text[:400]}")
    text = resp.json()["content"][0]["text"].strip()
    if text.startswith("```"):
        parts = text.split("```")
        text = parts[2] if len(parts) > 2 else parts[-1]
        text = text.lstrip("json").strip()
    # Fix truncated JSON: find last complete object/array
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        # Try to recover by finding the last valid closing brace
        for i in range(len(text)-1, 0, -1):
            if text[i] == '}':
                try:
                    return json.loads(text[:i+1])
                except json.JSONDecodeError:
                    continue
        raise RuntimeError("Could not parse Claude response. Try selecting fewer slip types or leagues.")

# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def build_excel(data):
    DK_GREEN = "1A5C2E"; MID_GREEN = "2E8B57"; LT_GREEN = "D6F0E0"
    ORANGE = "E67E22";   LT_ORANGE = "FEF0E0"; DARK_BLUE = "1A3A6E"
    LT_BLUE = "D6E8FA";  GOLD = "C8960C";      LT_GOLD = "FFF8DC"
    RED = "C0392B";      LT_RED = "FDECEA";    WHITE = "FFFFFF"
    DARK_GRAY = "2D2D2D"; MED_GRAY = "888888"

    TYPE_ACCENT = {
        "fortress": (DK_GREEN, LT_GREEN), "goals": (ORANGE, LT_ORANGE),
        "protection": (DARK_BLUE, LT_BLUE), "mixed": (MID_GREEN, LT_GREEN),
        "bonus": (GOLD, LT_GOLD), "insane": (RED, LT_RED),
    }

    def fill(c): return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color=DARK_GRAY, size=10, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    def aln(h="left", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
    def brd():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def banner(ws, row, text, bg, fg=WHITE, size=13, height=34, ncols=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=True, color=fg, size=size)
        c.fill = fill(bg); c.alignment = aln("center")
        ws.row_dimensions[row].height = height

    def hdr(ws, row, texts, bg=DK_GREEN):
        for col, txt in enumerate(texts, 1):
            c = ws.cell(row=row, column=col, value=txt)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
            c.fill = fill(bg); c.alignment = aln("center"); c.border = brd()
        ws.row_dimensions[row].height = 20

    def drow(ws, row, values, bg=WHITE, bold_cols=None, accent_col=None, accent_color=MID_GREEN):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            is_bold = bold_cols and col in bold_cols
            color = accent_color if accent_col and col == accent_col else DARK_GRAY
            c.font = Font(name="Arial", bold=is_bold, color=color, size=10)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center" if col in [1,3,5] else "left",
                                    vertical="center", wrap_text=True)
            c.border = brd()
        ws.row_dimensions[row].height = 22

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Overview"
    ws0.sheet_view.showGridLines = False
    for col, w in zip(range(1,8), [10, 28, 20, 14, 8, 14, 18]):
        ws0.column_dimensions[get_column_letter(col)].width = w

    banner(ws0, 1, f"⚽  BETTING SLIPS — {data.get('date','').upper()}  |  AI Generated",
           DK_GREEN, size=14, height=40, ncols=7)
    ws0.merge_cells("A2:G2")
    c2 = ws0.cell(row=2, column=1, value=data.get("summary",""))
    c2.font = Font(name="Arial", italic=True, color=WHITE, size=10)
    c2.fill = fill(MID_GREEN)
    c2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws0.row_dimensions[2].height = 28
    ws0.row_dimensions[3].height = 10

    slips = data.get("slips", [])
    hdr(ws0, 4, ["Slip","Title","Style","Est. Odds","Legs","Risk","Type"], ncols=7)
    RISK_BG = {"low": LT_GREEN, "medium": LT_GOLD, "high": LT_RED, "dream": LT_RED}
    for i, slip in enumerate(slips):
        r = 5 + i
        stype = slip.get("type","")
        risk_str = slip.get("risk","")
        rb = RISK_BG.get(risk_str.lower().split("-")[0], LT_GOLD)
        row_data = [f"Slip {i+1}", slip.get("title",""), slip.get("subtitle",""),
                    slip.get("estimatedOdds",""), slip.get("legs", len(slip.get("selections",[]))),
                    risk_str, stype.capitalize()]
        for col, val in enumerate(row_data, 1):
            c = ws0.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", bold=(stype=="insane"),
                          color=RED if stype=="insane" else DARK_GRAY, size=10)
            c.fill = fill(rb if col==6 else (LT_GREEN if i%2==0 else WHITE))
            c.alignment = Alignment(horizontal="left" if col in [2,3] else "center", vertical="center")
            c.border = brd()
        ws0.row_dimensions[r].height = 20

    fix_start = 6 + len(slips)
    ws0.row_dimensions[fix_start-1].height = 12
    banner(ws0, fix_start, "📋  ALL FIXTURES", DK_GREEN, size=12, height=26, ncols=7)
    hdr(ws0, fix_start+1, ["Time","League","Home","Home%","Away","Away%","Best Bet"], ncols=7)
    for j, f in enumerate(data.get("fixtures",[])):
        r = fix_start + 2 + j
        bg = LT_GREEN if j%2==0 else WHITE
        for col, val in enumerate([f.get("time",""), f"{f.get('flag','')} {f.get('league','')}",
                                    f.get("home",""), f"{f.get('homeWinPct','?')}%",
                                    f.get("away",""), f"{f.get('awayWinPct','?')}%",
                                    f.get("bestBet","")], 1):
            c = ws0.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=10, bold=(col==7),
                          color=MID_GREEN if col==7 else DARK_GRAY)
            c.fill = fill(bg)
            c.alignment = Alignment(horizontal="center" if col in [1,4,6] else "left", vertical="center")
            c.border = brd()
        ws0.row_dimensions[r].height = 18

    # Slip sheets
    for i, slip in enumerate(slips):
        stype = slip.get("type","")
        acc_main, acc_light = TYPE_ACCENT.get(stype, (MID_GREEN, LT_GREEN))
        icon = SLIP_TYPES.get(stype,("🎰",))[0]
        ws = wb.create_sheet(f"{icon} {stype.capitalize()[:12]}")
        ws.sheet_view.showGridLines = False
        for col, w in zip(range(1,6), [5, 32, 15, 34, 20]):
            ws.column_dimensions[get_column_letter(col)].width = w

        banner(ws, 1, f"{icon}  {slip.get('title','')}", acc_main, size=14, height=38)
        ws.merge_cells("A2:E2")
        c = ws.cell(row=2, column=1, value=slip.get("subtitle",""))
        c.font = Font(name="Arial", italic=True, color=WHITE, size=10)
        c.fill = fill(MID_GREEN); c.alignment = aln("center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 8

        for si, (lbl, val) in enumerate([("📈 Est. Odds", slip.get("estimatedOdds","?")),
                                          ("⚠️ Risk", slip.get("risk","?")),
                                          ("📋 Legs", slip.get("legs", len(slip.get("selections",[])))),
                                          ("📅 Date", data.get("date",""))], 1):
            cl = ws.cell(row=4, column=si, value=lbl)
            cl.font = Font(name="Arial", bold=True, color=WHITE, size=9)
            cl.fill = fill(acc_main); cl.alignment = aln("center"); cl.border = brd()
            cv = ws.cell(row=5, column=si, value=val)
            cv.font = Font(name="Arial", bold=True, color=DARK_GRAY, size=11)
            cv.fill = fill(acc_light); cv.alignment = aln("center"); cv.border = brd()
        ws.row_dimensions[4].height = 18; ws.row_dimensions[5].height = 24
        ws.row_dimensions[6].height = 8

        if stype == "insane":
            ws.merge_cells("A6:E6")
            cw = ws.cell(row=6, column=1,
                value="⚠️ STAKE €1-€5 ONLY — Dream slip. All realistic — but all landing together is rare!")
            cw.font = Font(name="Arial", bold=True, size=10, color="8B1A1A")
            cw.fill = fill("FFE4E1")
            cw.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cw.border = brd(); ws.row_dimensions[6].height = 28

        hdr(ws, 7, ["#","Match","League","Selection","Prob / Reasoning"])
        for j, s in enumerate(slip.get("selections",[])):
            r = 8 + j
            bg = acc_light if j%2==0 else WHITE
            drow(ws, r,
                [s.get("num",""), s.get("match",""),
                 f"{s.get('flag','')} {s.get('league','')}",
                 s.get("selection",""),
                 f"{s.get('prob','')} — {s.get('reasoning','')}"],
                bg=bg, bold_cols={4}, accent_col=4,
                accent_color=RED if stype=="insane" else MID_GREEN)

        last = 8 + len(slip.get("selections",[]))
        ws.row_dimensions[last].height = 8
        banner(ws, last+1, "🧠  ANALYSIS & REASONING", DK_GREEN, size=11, height=22)
        for k, line in enumerate(slip.get("analysis",[])):
            r = last + 2 + k
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            c = ws.cell(row=r, column=1, value=line)
            c.font = Font(name="Arial", size=10, color=DARK_GRAY)
            c.fill = fill(acc_light if k%2==0 else WHITE)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = brd(); ws.row_dimensions[r].height = 28

        disc = last + 2 + len(slip.get("analysis",[])) + 2
        ws.merge_cells(start_row=disc, start_column=1, end_row=disc, end_column=5)
        cd = ws.cell(row=disc, column=1,
            value="⚠️ For entertainment only. Gamble responsibly. Only bet what you can afford. 18+")
        cd.font = Font(name="Arial", italic=True, size=9, color=MED_GRAY)
        cd.alignment = aln("center"); ws.row_dimensions[disc].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
#  DISPLAY HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def risk_badge(risk):
    r = (risk or "").lower()
    if "low" in r:   return '<span class="tag tag-green">🟢 ' + risk + '</span>'
    if "dream" in r: return '<span class="tag tag-red">🔴 ' + risk + '</span>'
    if "high" in r:  return '<span class="tag tag-red">🔴 ' + risk + '</span>'
    return '<span class="tag tag-gold">🟡 ' + risk + '</span>'

def slip_badge(stype):
    css = SLIP_ACCENTS.get(stype, ("","","tag-green"))[2]
    icon = SLIP_TYPES.get(stype, ("🎰",))[0]
    name = SLIP_TYPES.get(stype, ("","?"))[1]
    return f'<span class="tag {css}">{icon} {name}</span>'

def render_slip(slip):
    stype = slip.get("type","")
    is_insane = stype == "insane"
    icon = SLIP_TYPES.get(stype, ("🎰",))[0]
    acc = SLIP_ACCENTS.get(stype, ("#2e8b57","#d6f0e0","tag-green"))
    header_class = "insane-header" if is_insane else "slip-header"
    header_style = f'background: linear-gradient(135deg, {acc[0]}, {acc[0]}cc);'

    if is_insane:
        header_html = f"""
        <div class="insane-header">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:1.6rem;letter-spacing:2px;margin-bottom:4px;">
                {icon} {slip.get('title','')}
            </div>
            <div style="font-size:0.88rem;opacity:0.85;margin-bottom:10px;">{slip.get('subtitle','')}</div>
            <div style="background:rgba(255,255,255,0.15);border-radius:6px;padding:8px 12px;font-size:0.85rem;font-weight:600;">
                ⚠️ STAKE €1–€5 ONLY — Dream slip. All realistic — but all 16 landing together is rare!
            </div>
        </div>"""
    else:
        header_html = f"""
        <div style="{header_style} border-radius:10px;padding:16px 20px;margin-bottom:16px;">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:1.5rem;letter-spacing:2px;margin-bottom:4px;">
                {icon} {slip.get('title','')}
            </div>
            <div style="font-size:0.88rem;opacity:0.85;">{slip.get('subtitle','')}</div>
        </div>"""

    st.markdown(header_html, unsafe_allow_html=True)

    # Stats row
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📈 Est. Odds", slip.get("estimatedOdds","?"))
    c2.metric("📋 Legs", slip.get("legs", len(slip.get("selections",[]))))
    c3.metric("⚠️ Risk", slip.get("risk","?"))
    c4.metric("🎯 Type", SLIP_TYPES.get(stype,("","?"))[1])

    st.markdown("---")

    # Selections
    st.markdown("#### 🎯 Selections")
    for s in slip.get("selections", []):
        num    = s.get("num","")
        match  = s.get("match","")
        league = f"{s.get('flag','')} {s.get('league','')}"
        sel    = s.get("selection","")
        prob   = s.get("prob","")
        reason = s.get("reasoning","")
        sel_color = "#ff8a7a" if is_insane else "#7ecf9e"

        st.markdown(f"""
        <div class="fixture-row">
            <div style="width:30px;height:30px;border-radius:50%;background:{acc[0]};
                        display:flex;align-items:center;justify-content:center;
                        font-size:0.8rem;font-weight:800;flex-shrink:0;margin-right:12px;">
                {num}
            </div>
            <div style="flex:1;min-width:0;">
                <div style="font-weight:600;color:#e8f0fe;font-size:0.9rem;">{match}</div>
                <div style="font-size:0.78rem;color:#7a8fa6;margin-top:2px;">{league} · {reason}</div>
            </div>
            <div style="font-weight:700;color:{sel_color};margin:0 16px;white-space:nowrap;font-size:0.9rem;">
                {sel}
            </div>
            <div style="font-weight:700;color:#f5a623;white-space:nowrap;font-size:0.9rem;">
                {prob}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # Analysis
    analysis = slip.get("analysis", [])
    if analysis:
        st.markdown("---")
        st.markdown("#### 🧠 Analysis")
        for line in analysis:
            st.markdown(f"""
            <div style="background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.06);
                        border-left:3px solid {acc[0]};border-radius:6px;
                        padding:10px 14px;margin-bottom:8px;font-size:0.9rem;color:#c8d8e8;line-height:1.5;">
                {line}
            </div>""", unsafe_allow_html=True)


def render_fixtures(fixtures):
    if not fixtures:
        st.info("No fixtures returned.")
        return
    for f in fixtures:
        home_pct  = f.get("homeWinPct","?")
        away_pct  = f.get("awayWinPct","?")
        draw_pct  = f.get("drawPct","?")
        best      = f.get("bestBet","")
        st.markdown(f"""
        <div class="fixture-row">
            <div style="width:55px;color:#7a8fa6;font-size:0.82rem;flex-shrink:0;">{f.get('time','')}</div>
            <div style="width:130px;font-size:0.82rem;color:#aaa;flex-shrink:0;">
                {f.get('flag','')} {f.get('league','')}
            </div>
            <div style="flex:1;font-weight:600;font-size:0.88rem;">
                {f.get('home','')} <span style="color:#7ecf9e;font-size:0.78rem;">({home_pct}%)</span>
            </div>
            <div style="width:30px;text-align:center;color:#7a8fa6;font-size:0.8rem;">vs</div>
            <div style="flex:1;font-weight:600;font-size:0.88rem;">
                {f.get('away','')} <span style="color:#aaa;font-size:0.78rem;">({away_pct}%)</span>
            </div>
            <div style="width:160px;font-weight:700;color:#7ecf9e;font-size:0.85rem;text-align:right;">
                {best}
            </div>
        </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:16px 0 8px;">
        <div style="font-size:2.5rem;">⚽</div>
        <div style="font-family:'Bebas Neue',sans-serif;font-size:1.6rem;letter-spacing:3px;color:#fff;">
            BetSlip AI
        </div>
        <div style="font-size:0.75rem;color:#7a8fa6;letter-spacing:1px;margin-top:2px;">
            POWERED BY CLAUDE
        </div>
    </div>
    <hr style="border-color:rgba(46,139,87,0.4);margin:12px 0;">
    """, unsafe_allow_html=True)

    # API Key — read from Streamlit secrets if available
    try:
        secret_key = st.secrets.get("ANTHROPIC_API_KEY", "")
    except Exception:
        secret_key = ""
    st.markdown("**🔑 Anthropic API Key**")
    if secret_key:
        st.success("✅ API key loaded from secrets")
        api_key = secret_key
    else:
        api_key = st.text_input("API Key", type="password", placeholder="sk-ant-...",
                                 help="Get yours at console.anthropic.com",
                                 key="api_key", label_visibility="collapsed")

    st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:16px 0;'>", unsafe_allow_html=True)

    # Date
    st.markdown("**📅 Match Date**")
    match_date = st.date_input("Match Date", value=datetime.date.today(), key="match_date",
                                label_visibility="collapsed")

    st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:16px 0;'>", unsafe_allow_html=True)

    # Leagues
    st.markdown("**🏆 Leagues**")
    selected_leagues = []
    league_defaults = ["bundesliga","la_liga","serie_a","ligue_1","eredivisie"]
    for lid, (flag, name) in LEAGUES.items():
        checked = st.checkbox(f"{flag} {name}", value=(lid in league_defaults), key=f"league_{lid}")
        if checked:
            selected_leagues.append(lid)

    st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:16px 0;'>", unsafe_allow_html=True)

    # Slip types
    st.markdown("**🎰 Slip Types**")
    selected_slips = []
    slip_defaults = ["fortress","goals","protection","mixed","bonus","insane"]
    for sid, (icon, name, desc) in SLIP_TYPES.items():
        checked = st.checkbox(f"{icon} {name}", value=(sid in slip_defaults), key=f"slip_{sid}",
                               help=desc)
        if checked:
            selected_slips.append(sid)

    st.markdown("<hr style='border-color:rgba(255,255,255,0.08);margin:16px 0;'>", unsafe_allow_html=True)

    generate_btn = st.button("🤖 GENERATE SLIPS", use_container_width=True)

    st.markdown("""
    <div style="margin-top:20px;padding:10px;background:rgba(192,57,43,0.15);
                border:1px solid rgba(192,57,43,0.3);border-radius:8px;
                font-size:0.75rem;color:#ff8a7a;text-align:center;">
        ⚠️ For entertainment only.<br>Gamble responsibly. 18+ only.
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  MAIN CONTENT
# ─────────────────────────────────────────────────────────────────────────────

# Hero header
st.markdown("""
<div style="text-align:center;padding:40px 0 20px;">
    <div style="font-family:'Bebas Neue',sans-serif;font-size:3.5rem;letter-spacing:4px;
                background:linear-gradient(135deg,#2e8b57,#f5a623);
                -webkit-background-clip:text;-webkit-text-fill-color:transparent;
                line-height:1.1;">
        BETSLIP AI GENERATOR
    </div>
    <div style="color:#7a8fa6;font-size:1rem;margin-top:8px;letter-spacing:2px;">
        🇩🇪 BUNDESLIGA &nbsp;·&nbsp; 🇪🇸 LA LIGA &nbsp;·&nbsp; 🇮🇹 SERIE A &nbsp;·&nbsp;
        🇫🇷 LIGUE 1 &nbsp;·&nbsp; 🇳🇱 EREDIVISIE &nbsp;·&nbsp; 🏴 EPL &nbsp;·&nbsp; 🏆 UCL
    </div>
</div>
""", unsafe_allow_html=True)

# State
if "result" not in st.session_state:
    st.session_state.result = None
if "error" not in st.session_state:
    st.session_state.error = None

# ── GENERATE ─────────────────────────────────────────────────────────────────
if generate_btn:
    if not api_key:
        st.error("⚠️ Please enter your Anthropic API key in the sidebar.")
    elif not selected_leagues:
        st.error("⚠️ Select at least one league.")
    elif not selected_slips:
        st.error("⚠️ Select at least one slip type.")
    else:
        date_str = match_date.strftime("%A %d %B %Y")
        with st.spinner(f"⏳ Claude is analysing fixtures for {date_str}... (10–25 seconds)"):
            try:
                result = call_claude(api_key, date_str, selected_leagues, selected_slips)
                st.session_state.result = result
                st.session_state.error  = None
            except Exception as e:
                st.session_state.error  = str(e)
                st.session_state.result = None

# ── ERROR ─────────────────────────────────────────────────────────────────────
if st.session_state.error:
    st.error(f"❌ {st.session_state.error}")

# ── RESULTS ───────────────────────────────────────────────────────────────────
if st.session_state.result:
    data = st.session_state.result
    slips = data.get("slips", [])
    fixtures = data.get("fixtures", [])

    # Summary banner
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1a5c2e,#2e8b57);border-radius:14px;
                padding:20px 24px;margin:8px 0 24px;border:1px solid rgba(245,166,35,0.4);">
        <div style="font-family:'Bebas Neue',sans-serif;font-size:1.8rem;letter-spacing:2px;margin-bottom:6px;">
            📅 {data.get('date','')}
        </div>
        <div style="font-size:0.95rem;color:rgba(255,255,255,0.85);line-height:1.6;">
            {data.get('summary','')}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Stats row
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🎰 Slips Generated", len(slips))
    c2.metric("📋 Fixtures", len(fixtures))
    c3.metric("🏆 Leagues", len(selected_leagues))
    c4.metric("📅 Date", match_date.strftime("%d %b %Y"))

    st.markdown("---")

    # Excel download
    try:
        xl_bytes = build_excel(data)
        fname = f"BettingSlips_{match_date.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            label="📥 DOWNLOAD EXCEL FILE",
            data=xl_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as e:
        st.warning(f"Excel export unavailable: {e}")

    st.markdown("---")

    # Tabs: slips + fixtures
    slip_labels = [f"{SLIP_TYPES.get(s.get('type',''),('🎰',))[0]} {s.get('title','Slip')[:18]}" for s in slips]
    all_tabs = slip_labels + ["📋 All Fixtures"]
    tabs = st.tabs(all_tabs)

    for i, (tab, slip) in enumerate(zip(tabs[:-1], slips)):
        with tab:
            render_slip(slip)

    with tabs[-1]:
        st.markdown("#### 📋 All Fixtures")
        render_fixtures(fixtures)

else:
    # Empty state
    st.markdown("""
    <div style="text-align:center;padding:60px 20px;color:#7a8fa6;">
        <div style="font-size:4rem;margin-bottom:16px;">⚽</div>
        <div style="font-family:'Bebas Neue',sans-serif;font-size:1.8rem;letter-spacing:2px;color:#3d5a40;margin-bottom:10px;">
            READY TO GENERATE
        </div>
        <div style="font-size:0.95rem;line-height:1.7;max-width:420px;margin:0 auto;">
            Enter your Anthropic API key, pick a date and leagues in the sidebar,
            then hit <strong style="color:#2e8b57;">Generate Slips</strong> to get
            AI-powered betting analysis across all major European leagues.
        </div>
    </div>
    """, unsafe_allow_html=True)

    # How it works
    st.markdown("---")
    st.markdown("### 🤖 How It Works")
    col1, col2, col3, col4 = st.columns(4)
    for col, num, title, desc in [
        (col1,"1️⃣","Pick a Date","Choose any upcoming match day"),
        (col2,"2️⃣","Select Leagues","Bundesliga, La Liga, Serie A, Ligue 1, Eredivisie & more"),
        (col3,"3️⃣","Choose Slip Types","Fortress, Goals, Protection, Insane & more"),
        (col4,"4️⃣","Download Excel","Get full formatted slips + all fixtures"),
    ]:
        with col:
            st.markdown(f"""
            <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);
                        border-radius:12px;padding:20px;text-align:center;">
                <div style="font-size:1.8rem;margin-bottom:8px;">{num}</div>
                <div style="font-family:'Bebas Neue',sans-serif;font-size:1.1rem;letter-spacing:1px;
                            color:#2e8b57;margin-bottom:6px;">{title}</div>
                <div style="font-size:0.82rem;color:#7a8fa6;line-height:1.5;">{desc}</div>
            </div>""", unsafe_allow_html=True)
